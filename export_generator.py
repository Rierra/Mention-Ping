#!/usr/bin/env python3
"""
Export Generator Module
Generates CSV and XLSX export files from mention history data.
"""

import csv
import os
import io
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# Column definitions for export
EXPORT_COLUMNS = [
    'Date',
    'Type',
    'Subreddit',
    'Author',
    'Title',
    'Content',
    'Keyword Matched',
    'URL',
    'Upvotes',
    'Comment Count',
    'Parent Post ID'
]


def filter_by_date(
    records: List[Dict],
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None
) -> List[Dict]:
    """Filter records by date range.
    
    Args:
        records: List of mention records
        start_date: Start of date range (inclusive)
        end_date: End of date range (inclusive)
    
    Returns:
        Filtered list of records
    """
    if not start_date and not end_date:
        return records
    
    filtered = []
    for record in records:
        try:
            record_date = datetime.fromisoformat(record.get('date', ''))
            
            if start_date and record_date < start_date:
                continue
            if end_date and record_date > end_date:
                continue
            
            filtered.append(record)
        except (ValueError, TypeError):
            # Skip records with invalid dates
            continue
    
    return filtered


def filter_by_preset(records: List[Dict], preset: str) -> List[Dict]:
    """Filter records by preset time period.
    
    Args:
        records: List of mention records
        preset: One of 'day', 'week', 'month'
    
    Returns:
        Filtered list of records
    """
    now = datetime.now()
    
    if preset == 'day':
        start_date = now - timedelta(days=1)
    elif preset == 'week':
        start_date = now - timedelta(weeks=1)
    elif preset == 'month':
        start_date = now - timedelta(days=30)
    else:
        return records
    
    return filter_by_date(records, start_date=start_date)


def generate_csv(records: List[Dict], filepath: str) -> int:
    """Generate CSV file from records.
    
    Args:
        records: List of mention records
        filepath: Output file path
    
    Returns:
        Number of rows written
    """
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS, extrasaction='ignore')
        writer.writeheader()
        
        for record in records:
            # Map internal keys to column names
            row = {
                'Date': record.get('date', ''),
                'Type': record.get('type', ''),
                'Subreddit': record.get('subreddit', ''),
                'Author': record.get('author', ''),
                'Title': record.get('title', ''),
                'Content': record.get('content', ''),
                'Keyword Matched': record.get('keyword_matched', ''),
                'URL': record.get('url', ''),
                'Upvotes': record.get('upvotes', ''),
                'Comment Count': record.get('comment_count', ''),
                'Parent Post ID': record.get('parent_post_id', '')
            }
            writer.writerow(row)
    
    return len(records)


def generate_csv_bytes(records: List[Dict]) -> bytes:
    """Generate CSV as bytes (for sending without saving to disk).
    
    Args:
        records: List of mention records
    
    Returns:
        CSV content as bytes
    """
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS, extrasaction='ignore')
    writer.writeheader()
    
    for record in records:
        row = {
            'Date': record.get('date', ''),
            'Type': record.get('type', ''),
            'Subreddit': record.get('subreddit', ''),
            'Author': record.get('author', ''),
            'Title': record.get('title', ''),
            'Content': record.get('content', ''),
            'Keyword Matched': record.get('keyword_matched', ''),
            'URL': record.get('url', ''),
            'Upvotes': record.get('upvotes', ''),
            'Comment Count': record.get('comment_count', ''),
            'Parent Post ID': record.get('parent_post_id', '')
        }
        writer.writerow(row)
    
    return output.getvalue().encode('utf-8')


def generate_xlsx(records: List[Dict], filepath: str) -> int:
    """Generate Excel file from records.
    
    Args:
        records: List of mention records
        filepath: Output file path
    
    Returns:
        Number of rows written
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mentions"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Write headers
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record.get('date', ''))
        ws.cell(row=row_idx, column=2, value=record.get('type', ''))
        ws.cell(row=row_idx, column=3, value=record.get('subreddit', ''))
        ws.cell(row=row_idx, column=4, value=record.get('author', ''))
        ws.cell(row=row_idx, column=5, value=record.get('title', ''))
        ws.cell(row=row_idx, column=6, value=record.get('content', ''))
        ws.cell(row=row_idx, column=7, value=record.get('keyword_matched', ''))
        ws.cell(row=row_idx, column=8, value=record.get('url', ''))
        ws.cell(row=row_idx, column=9, value=record.get('upvotes', ''))
        ws.cell(row=row_idx, column=10, value=record.get('comment_count', ''))
        ws.cell(row=row_idx, column=11, value=record.get('parent_post_id', ''))
    
    # Auto-adjust column widths (with max width)
    for col_idx, _ in enumerate(EXPORT_COLUMNS, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        # Cap width at 50 characters
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(filepath)
    return len(records)


def generate_xlsx_bytes(records: List[Dict]) -> bytes:
    """Generate Excel as bytes (for sending without saving to disk).
    
    Args:
        records: List of mention records
    
    Returns:
        Excel content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mentions"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Write headers
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record.get('date', ''))
        ws.cell(row=row_idx, column=2, value=record.get('type', ''))
        ws.cell(row=row_idx, column=3, value=record.get('subreddit', ''))
        ws.cell(row=row_idx, column=4, value=record.get('author', ''))
        ws.cell(row=row_idx, column=5, value=record.get('title', ''))
        ws.cell(row=row_idx, column=6, value=record.get('content', ''))
        ws.cell(row=row_idx, column=7, value=record.get('keyword_matched', ''))
        ws.cell(row=row_idx, column=8, value=record.get('url', ''))
        ws.cell(row=row_idx, column=9, value=record.get('upvotes', ''))
        ws.cell(row=row_idx, column=10, value=record.get('comment_count', ''))
        ws.cell(row=row_idx, column=11, value=record.get('parent_post_id', ''))
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_file_size_str(size_bytes: int) -> str:
    """Convert bytes to human-readable size string.
    
    Args:
        size_bytes: Size in bytes
    
    Returns:
        Human-readable size string (e.g., "124 KB")
    """
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"


def parse_date_arg(date_str: str) -> Optional[datetime]:
    """Parse date string in YYYY-MM-DD format.
    
    Args:
        date_str: Date string
    
    Returns:
        datetime object or None if parsing fails
    """
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return None


def get_export_stats(records: List[Dict]) -> Dict:
    """Calculate statistics for export.
    
    Args:
        records: List of mention records
    
    Returns:
        Dictionary with stats
    """
    mentions = sum(1 for r in records if r.get('type') in ('post', 'comment'))
    context = sum(1 for r in records if r.get('type') == 'context_comment')
    
    dates = []
    for r in records:
        try:
            dates.append(datetime.fromisoformat(r.get('date', '')))
        except:
            pass
    
    return {
        'total_rows': len(records),
        'mentions': mentions,
        'context_comments': context,
        'start_date': min(dates).strftime('%b %d, %Y') if dates else 'N/A',
        'end_date': max(dates).strftime('%b %d, %Y') if dates else 'N/A'
    }
